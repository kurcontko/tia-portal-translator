"""Tests for the Excel writer."""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from tia_portal_translator.config import Config
from tia_portal_translator.writers import ExcelWriter


def test_excel_writer_atomic_save_same_path(tmp_path: Path):
    """Ensure saving to the same path uses atomic replace safely."""
    path = tmp_path / "input.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "User Texts"
    sheet["A1"] = "en-US"
    sheet["B1"] = "de-DE"
    sheet["A2"] = "hello"
    workbook.save(path)

    workbook = load_workbook(path)
    sheet = workbook["User Texts"]

    config = Config(excel_file=str(path), output_file=str(path))
    writer = ExcelWriter(config, workbook, sheet)
    writer.write_translations([(2, "hallo")], "B")
    writer.save_workbook()

    reloaded = load_workbook(path)
    assert reloaded["User Texts"]["B2"].value == "hallo"
