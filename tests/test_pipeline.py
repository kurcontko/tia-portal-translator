"""Tests for the translation pipeline."""

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from conftest import (
    FailingBatchService,
    LongBatchService,
    RecordingTranslationService,
    ShortBatchService,
)
from tia_portal_translator.config import Config
from tia_portal_translator.pipeline import TranslatorPipeline


@pytest.mark.asyncio
async def test_pipeline_translates_in_chunks(tmp_path: Path, sample_workbook):
    """Test that pipeline correctly processes texts in chunks."""
    source_path = sample_workbook("en-US", "de-DE", ["one", "two", "", "four"], "input.xlsx")
    output_path = tmp_path / "output.xlsx"

    config = Config(
        excel_file=str(source_path),
        output_file=str(output_path),
        chunk_size=2,
    )
    service = RecordingTranslationService(suffix="-de")
    pipeline = TranslatorPipeline(config, service)

    await pipeline.translate_project("en-US", "de-DE")

    workbook = load_workbook(output_path)
    sheet = workbook["User Texts"]
    assert sheet["B2"].value == "one-de"
    assert sheet["B3"].value == "two-de"
    assert sheet["B4"].value is None  # Empty cells in Excel are None, not ""
    assert sheet["B5"].value == "four-de"
    assert service.batch_sizes == [2, 2]


@pytest.mark.asyncio
async def test_pipeline_handles_chunk_failure(tmp_path: Path, sample_workbook):
    """Test that pipeline handles translation failures gracefully."""
    source_path = sample_workbook("en-US", "de-DE", ["one", "two"], "input.xlsx")
    output_path = tmp_path / "output.xlsx"

    config = Config(
        excel_file=str(source_path),
        output_file=str(output_path),
        chunk_size=10,
    )
    pipeline = TranslatorPipeline(config, FailingBatchService())

    await pipeline.translate_project("en-US", "de-DE")

    workbook = load_workbook(output_path)
    sheet = workbook["User Texts"]
    assert sheet["B2"].value is None  # Failed translation results in empty cell
    assert sheet["B3"].value is None


@pytest.mark.asyncio
async def test_pipeline_fail_fast_raises(tmp_path: Path, sample_workbook):
    """Test that fail-fast aborts on translation errors."""
    source_path = sample_workbook("en-US", "de-DE", ["one", "two"], "input.xlsx")
    output_path = tmp_path / "output.xlsx"

    config = Config(
        excel_file=str(source_path),
        output_file=str(output_path),
        chunk_size=10,
        fail_fast=True,
    )
    pipeline = TranslatorPipeline(config, FailingBatchService())

    with pytest.raises(RuntimeError):
        await pipeline.translate_project("en-US", "de-DE")


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("service_cls", "expected_count", "actual_count", "provider_name"),
    [
        (ShortBatchService, 3, 2, "short-batch"),
        (LongBatchService, 3, 4, "long-batch"),
    ],
)
async def test_pipeline_handles_batch_size_mismatch(
    tmp_path: Path,
    sample_workbook,
    service_cls,
    expected_count,
    actual_count,
    provider_name,
):
    """Test that pipeline handles short/long batch responses with errors."""
    source_path = sample_workbook("en-US", "de-DE", ["one", "two", "three"], "input.xlsx")
    output_path = tmp_path / "output.xlsx"
    report_path = tmp_path / "report.json"

    config = Config(
        excel_file=str(source_path),
        output_file=str(output_path),
        chunk_size=10,
        report_path=str(report_path),
    )
    pipeline = TranslatorPipeline(config, service_cls())

    await pipeline.translate_project("en-US", "de-DE")

    workbook = load_workbook(output_path)
    sheet = workbook["User Texts"]
    assert sheet["B2"].value is None
    assert sheet["B3"].value is None
    assert sheet["B4"].value is None

    data = json.loads(report_path.read_text(encoding="utf-8"))
    assert len(data) == 3
    expected_message = f"expected {expected_count} items, got {actual_count}"
    for entry in data:
        assert entry["error"]
        assert provider_name in entry["error"]
        assert expected_message in entry["error"]


@pytest.mark.asyncio
async def test_pipeline_writes_report(tmp_path: Path, sample_workbook):
    """Test that pipeline emits a JSON report when requested."""
    source_path = sample_workbook("en-US", "de-DE", ["one", "two"], "input.xlsx")
    output_path = tmp_path / "output.xlsx"
    report_path = tmp_path / "report.json"

    config = Config(
        excel_file=str(source_path),
        output_file=str(output_path),
        chunk_size=2,
        report_path=str(report_path),
    )
    pipeline = TranslatorPipeline(config, RecordingTranslationService(suffix="-de"))

    await pipeline.translate_project("en-US", "de-DE")

    data = json.loads(report_path.read_text(encoding="utf-8"))
    assert len(data) == 2
    assert data[0]["row_num"] == 2
    assert data[0]["translated_text"] == "one-de"


@pytest.mark.asyncio
async def test_pipeline_item_exception_records_error(tmp_path: Path, sample_workbook):
    """Test that per-item failures are captured in the report."""
    source_path = sample_workbook("en-US", "de-DE", ["one", "two"], "input.xlsx")
    output_path = tmp_path / "output.xlsx"
    report_path = tmp_path / "report.json"

    class ItemErrorService:
        service_name = "item-error"

        async def translate_batch(self, texts):
            return [f"{texts[0]}-de", RuntimeError("boom")]

    config = Config(
        excel_file=str(source_path),
        output_file=str(output_path),
        chunk_size=10,
        report_path=str(report_path),
    )
    pipeline = TranslatorPipeline(config, ItemErrorService())

    await pipeline.translate_project("en-US", "de-DE")

    workbook = load_workbook(output_path)
    sheet = workbook["User Texts"]
    assert sheet["B2"].value == "one-de"
    assert sheet["B3"].value is None

    data = json.loads(report_path.read_text(encoding="utf-8"))
    assert data[1]["error"]
    assert "item_translate_error" in data[1]["error"]


@pytest.mark.asyncio
async def test_pipeline_invalid_report_extension_fail_fast(tmp_path: Path, sample_workbook):
    """Test that invalid report extensions raise when fail_fast is enabled."""
    source_path = sample_workbook("en-US", "de-DE", ["one"], "input.xlsx")
    output_path = tmp_path / "output.xlsx"
    report_path = tmp_path / "report.txt"

    config = Config(
        excel_file=str(source_path),
        output_file=str(output_path),
        chunk_size=10,
        report_path=str(report_path),
        fail_fast=True,
    )
    pipeline = TranslatorPipeline(config, RecordingTranslationService(suffix="-de"))

    with pytest.raises(ValueError, match="Report path must end with .json or .csv"):
        await pipeline.translate_project("en-US", "de-DE")


@pytest.mark.asyncio
async def test_pipeline_missing_target_column_raises(tmp_path: Path, sample_workbook):
    """Test that missing target columns raise clear errors."""
    source_path = sample_workbook("en-US", "fr-FR", ["one"], "input.xlsx")
    output_path = tmp_path / "output.xlsx"

    config = Config(
        excel_file=str(source_path),
        output_file=str(output_path),
        chunk_size=10,
    )
    pipeline = TranslatorPipeline(config, RecordingTranslationService(suffix="-de"))

    with pytest.raises(ValueError, match="Target column 'de-DE' not found"):
        await pipeline.translate_project("en-US", "de-DE")
