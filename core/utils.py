# core/utils.py
from openpyxl import load_workbook
from googletrans import Translator
import openai
from deepl import Translator as DeepLTranslator

def translation_service_factory(service, api_key=None, destination_language=None):
    if service == 'google':
        return GoogleTranslationService(api_key, destination_language)
    elif service == 'gpt':
        return GPTTranslationService(api_key, destination_language)
    elif service == 'deepl':
        return DeepLTranslationService(api_key, destination_language)
    else:
        raise ValueError(f'Invalid service: {service}')
    
class TranslationService:
    def __init__(self, api_key=None, destination_language=None):
        self.api_key = api_key
        self.destination_language = destination_language

    def translate(self, text):
        pass

class GoogleTranslationService(TranslationService):
    def translate(self, text):
        translator = Translator()
        return translator.translate(text, dest=self.destination_language).text

class GPTTranslationService(TranslationService):
    def __init__(self, api_key=None, destination_language=None):
        super().__init__(api_key, destination_language)
        self.delimiter = ' ||| '  # delimiter to separate cells in a prompt

    def translate(self, cells):
        texts = [cell.value for cell in cells]
        prompt = f'Translate the following texts to "{self.destination_language}" language:'
        for text in texts:
            prompt += f'\n{text}{self.delimiter}'

        openai.api_key = self.api_key
        response = openai.Completion.create(
            engine='text-davinci-002',
            prompt=prompt,
            max_tokens=300,  # increased max_tokens due to multiple cells
            n=1,
            stop=None,
            temperature=0.5,
        )

        # Parse the response
        response_texts = response.choices[0].text.strip().split(self.delimiter)
        return response_texts
        
class DeepLTranslationService(TranslationService):
    def translate(self, text):
        translator = DeepLTranslator(self.api_key)
        return translator.translate_text(text, target_lang=self.destination_language)
    
def process_frame_google(chunk_tuple, translator_instance, ws, destination_to_translation_col):
    index, chunk = chunk_tuple
    print(f'Translating chunk {index+1}...')
    translated_chunk = [(translator_instance.translate(cell.value) if cell.value else ws[destination_to_translation_col][index].value) for cell in chunk]
    return index, translated_chunk

def process_frame_gpt(chunk_tuple, translator_instance, ws, destination_to_translation_col):
    index, chunk = chunk_tuple
    print(f'Translating chunk {index+1}...')
    translated_chunk = translator_instance.translate(chunk)
    return index, translated_chunk

def find_column_letter(column_name, ws):
    for cell in ws[1]:
        if cell.value == column_name:
            return cell.column_letter
    return None