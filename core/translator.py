# core/translator.py
import os
import multiprocessing as mp
from openpyxl import load_workbook

from .utils import find_column_letter, translation_service_factory, process_frame_google, process_frame_gpt

n_processes = min(os.cpu_count(), 1)

def translate_excel(translator_service, source_to_translation, destination_to_translation, my_excel='TIAProjectTexts.xlsx', my_excel_sheet_name='User Texts'):
    
    result_excel = f'{my_excel[:-5]}_translated.xlsx'

    try:
        translator_service = translator_service
        source_to_translation = source_to_translation
        destination_to_translation = destination_to_translation

        # Extract the destination language from the column name
        destination_language = destination_to_translation.split('-')[0]

        # Check if the API key is required and available
        api_key = None
        if translator_service == 'gpt' or translator_service == 'deepl':
            api_key_env_var = 'OPENAI_API_KEY' if translator_service == 'gpt' else 'DEEPL_API_KEY'
            try:
                api_key = os.environ[api_key_env_var]
            except KeyError:
                print(f'Error: {translator_service.upper()} translation requires the {api_key_env_var} environment variable.')
                exit(1)

        # Read Excel file
        wb = load_workbook(my_excel)
        ws = wb[my_excel_sheet_name]

        # Find column letters for source and destination
        source_to_translation_col = find_column_letter(source_to_translation, ws)
        destination_to_translation_col = find_column_letter(destination_to_translation, ws)

        if not source_to_translation_col or not destination_to_translation_col:
            print('Could not find column names.')
            exit(1)

        # Split data into chunks
        row_count = ws.max_row - 1
        chunk_size = row_count // n_processes
        data_chunks = [(i, ws[source_to_translation_col][i*chunk_size+1:(i+1)*chunk_size+1]) for i in range(n_processes)]

        # Instantiate the translator
        translator_instance = translation_service_factory(translator_service, api_key, destination_language)

        process_frame_map = {
            'google': process_frame_google,
            'gpt': process_frame_gpt,
            'deepl': process_frame_google  # DeepL uses the same technique as Google for now
        }

        # Retrieve the appropriate process_frame function for the current service
        process_frame = process_frame_map[translator_service]

        # Use multiprocessing to translate chunks
        pool = mp.Pool(n_processes)
        result_list = pool.starmap(process_frame, [(chunk_tuple, translator_instance) for chunk_tuple in data_chunks], ws , destination_to_translation_col)
        pool.close()
        pool.join()

        # Sort results by index
        result_list.sort(key=lambda x: x[0])

        # Write translations back to the worksheet
        for index, chunk in result_list:
            if translator_service == 'gpt':
                for idx, cell in enumerate(chunk):
                    ws[f'{destination_to_translation_col}{index * chunk_size + idx + 2}'].value = cell.strip()  # strip() is used to remove leading and trailing whitespace
            else:
                for idx, cell in enumerate(chunk):
                    ws[f'{destination_to_translation_col}{index * chunk_size + idx + 2}'].value = cell

        # Save the workbook
        if os.path.exists(result_excel):
            print(f'Removed file {result_excel}')
        wb.save(result_excel)
        print(f'Created new file {result_excel}')
        print('Translating finished!')

    except Exception as e:
        print(f'An error occurred: {e}')
