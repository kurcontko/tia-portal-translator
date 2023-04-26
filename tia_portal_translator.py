import os
import multiprocessing as mp
import argparse
from openpyxl import load_workbook
from googletrans import Translator  # use 3.1.0a0 or later
import openai
from deepl import Translator as DeepLTranslator

my_excel = 'IG11-Texts.xlsx'
my_excel_sheet_name = 'User Texts'
n_processes = min(os.cpu_count(), 64) #64 is maximum number in Windows, you can try to push the no of processes to the limits, but it can hit your system's stability
result_excel = f'{my_excel[:-5]}_translated.xlsx'
source_to_translation = 'en-US' 
destination_to_translation = 'pt-PT'
# Extract the destination language from the column name
destination_language = destination_to_translation.split('-')[0]

def translate_with_chatgpt(text, api_key, dest_language):
    openai.api_key = api_key

    prompt = f'Translate the following text to {dest_language}:\n{text}'
    response = openai.Completion.create(
        engine='text-davinci-002',
        prompt=prompt,
        max_tokens=100,
        n=1,
        stop=None,
        temperature=0.5,
    )

    translated_text = response.choices[0].text.strip()
    return translated_text

def translate_with_deepl(text, api_key, dest_language):
    translator = DeepLTranslator(api_key)
    translated_text = translator.translate_text(text, target_lang=dest_language)
    return translated_text

def process_frame(chunk_tuple, translator_service, api_key=None):
    index, chunk = chunk_tuple

    print(f'Translating chunk {index+1}...')
    translated_chunk = []

    if translator_service == 'google':
        translator = Translator()
        translated_chunk = [(translator.translate(cell.value, dest=destination_language).text if cell.value else "") for cell in chunk]
    elif translator_service == 'gpt':
        for cell in chunk:
            if cell.value:
                translated_text = translate_with_chatgpt(cell.value, api_key, destination_language)
                translated_chunk.append(translated_text)
            else:
                translated_chunk.append("")
    elif translator_service == 'deepl':
        for cell in chunk:
            if cell.value:
                translated_text = translate_with_deepl(cell.value, api_key, destination_language)
                translated_chunk.append(translated_text)
            else:
                translated_chunk.append("")

    return index, translated_chunk

def find_column_letter(column_name, ws):
    for cell in ws[1]:
        if cell.value == column_name:
            return cell.column_letter
    return None

def parse_arguments():
    parser = argparse.ArgumentParser(description='Translate an Excel file using Google Translate, GPT, or DeepL.')
    parser.add_argument('--service', choices=['google', 'gpt', 'deepl'], required=True, help='Choose the translation service (google, gpt, or deepl).')
    args = parser.parse_args()
    return args

if __name__ == '__main__':
    try:

        # Parse command-line arguments
        args = parse_arguments()
        translator_service = args.service

       # Check if the API key is required and available
        api_key = None
        if translator_service == 'gpt' or translator_service == 'deepl':
            api_key_env_var = 'OPENAI_API_KEY' if translator_service == 'gpt' else 'DEEPL_API_KEY'
            try:
                api_key = os.environ[api_key_env_var]
            except KeyError:
                print(f'Error: {translator_service.upper()} translation requires the {api_key_env_var} environment variable.')
                exit(1)

        # Read Excel file
        wb = load_workbook(my_excel)
        ws = wb[my_excel_sheet_name]

        # Find column letters for source and destination
        source_to_translation_col = find_column_letter(source_to_translation, ws)
        destination_to_translation_col = find_column_letter(destination_to_translation, ws)

        if not source_to_translation_col or not destination_to_translation_col:
            print('Could not find column names.')
            exit(1)

        # Split data into chunks
        row_count = ws.max_row - 1
        chunk_size = row_count // n_processes
        data_chunks = [(i, ws[source_to_translation_col][i*chunk_size+1:(i+1)*chunk_size+1]) for i in range(n_processes)]

        # Use multiprocessing to translate chunks
        pool = mp.Pool(n_processes)
        result_list = pool.starmap(process_frame, [(chunk_tuple, translator_service, api_key) for chunk_tuple in data_chunks])
        pool.close()
        pool.join()

        # Sort results by index
        result_list.sort(key=lambda x: x[0])

        # Write translations back to the worksheet
        for index, chunk in result_list:
            for idx, cell in enumerate(chunk):
                ws[f'{destination_to_translation_col}{index * chunk_size + idx + 2}'].value = cell

        # Save the workbook
        if os.path.exists(result_excel):
            print(f'Removed file {result_excel}')
        wb.save(result_excel)
        print(f'Created new file {result_excel}')
        print('Translating finished!')
    except Exception as e:
        print(f'An error occurred: {e}')
