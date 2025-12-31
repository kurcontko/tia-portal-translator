import logging
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from tia_portal_translator.config import Config

logger = logging.getLogger(__name__)


class ExcelReader:
    """Read data from the Excel source workbook."""

    def __init__(self, config: Config) -> None:
        self.config = config
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None

    def load_workbook(self) -> None:
        """Load the Excel workbook."""
        try:
            self.workbook = load_workbook(self.config.excel_file)
            self.worksheet = self.workbook[self.config.sheet_name]
            logger.info("Loaded workbook: %s", self.config.excel_file)
        except FileNotFoundError as exc:
            raise FileNotFoundError(f"Excel file not found: {self.config.excel_file}") from exc
        except KeyError as exc:
            raise KeyError(f"Sheet '{self.config.sheet_name}' not found in workbook") from exc

    def find_column_letter(self, column_name: str) -> Optional[str]:
        """Find the column letter for a given column name."""
        if not self.worksheet:
            raise ValueError("Workbook not loaded")

        for cell in self.worksheet[1]:
            if cell.value == column_name:
                return cell.column_letter
        return None

    def get_source_texts(self, source_column: str) -> list[tuple[int, str]]:
        """Get source texts from the specified column."""
        if not self.worksheet:
            raise ValueError("Workbook not loaded")

        texts = []
        for row_num, cell in enumerate(self.worksheet[source_column][1:], start=2):
            if self.config.skip_formulas:
                if cell.data_type == "f":
                    texts.append((row_num, ""))
                    continue
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    texts.append((row_num, ""))
                    continue

            if cell.value is None:
                texts.append((row_num, ""))
            else:
                texts.append((row_num, str(cell.value)))

        return texts
